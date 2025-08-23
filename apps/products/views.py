import io
import json
import traceback
from openpyxl.styles import Font
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from django.utils import timezone
from django.db.models import Sum, Value
from django.db.models.functions import Coalesce
from apps.prices.models import Price
from apps.products.models import Product, ProductCategory
from apps.products.serializers import ProductCategorySerializer, ProductSerializer
from apps.units_of_measurement.models import UnitOfMeasurement
from django.db import models
from utils.parse_excel import parse_excel
from utils.parse_csv import parse_csv

class ProductsView(APIView):
  permission_classes = [IsAuthenticated]
  parser_classes = [JSONParser, FormParser, MultiPartParser]

  def get(self, request, format=None):
    products = Product.objects.filter(deleted_at__isnull=True) \
      .prefetch_related('categories') \
      .annotate(
        total_stock=Coalesce(Sum('stock_controls__current_stock'), Value(0)),
        total_booking=Coalesce(Sum('stock_controls__current_booking'), Value(0))
      )
    serializer = ProductSerializer(products, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)
  
class ProductDetailView(APIView):
  permission_classes = [IsAuthenticated]
  parser_classes = [JSONParser, FormParser, MultiPartParser]

  def get(self, request, pk, format=None):
    product = get_object_or_404(Product.objects.prefetch_related('categories'), pk=pk, deleted_at__isnull=True)
    serializer = ProductSerializer(product)
    return Response(serializer.data, status=status.HTTP_200_OK)
  
class CreateProductView(APIView):
  permission_classes = [IsAuthenticated]
  parser_classes = [JSONParser, FormParser, MultiPartParser]

  def post(self, request, format=None):
    serializer = ProductSerializer(data=request.data)

    if serializer.is_valid():
      new_product = serializer.save(created_by=request.user)

      # Crear registro en prices automáticamente al crear producto
      Price.objects.create(
        product=new_product,
        cost=new_product.cost,
        price_cf=new_product.featured_pcf,
        price_sf=new_product.featured_psf,
        price_box=new_product.featured_pbox,
        created_by=request.user,
        updated_by=request.user,
      )

      return Response(serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UpdateProductView(APIView):
  permission_classes = [IsAuthenticated]
  parser_classes = [JSONParser, FormParser, MultiPartParser]

  def put(self, request, pk, format=None):
    product = get_object_or_404(Product, pk=pk, deleted_at__isnull=True)
    old_cost = product.cost  # Guardamos el costo antes de la actualización

    serializer = ProductSerializer(product, data=request.data, partial=True)

    if serializer.is_valid():
      updated_product = serializer.save(
        updated_by=request.user,
        updated_at=timezone.now()
      )

      # Check if cost was modified
      new_cost = updated_product.cost
      if "cost" in request.data and str(old_cost) != str(new_cost):
        Price.objects.create(
          product=updated_product,
          cost=new_cost,
          price_cf=updated_product.featured_pcf,
          price_sf=updated_product.featured_psf,
          price_box=updated_product.featured_pbox,
          created_by=request.user,
          updated_by=request.user,
        )

      return Response(serializer.data, status=status.HTTP_200_OK)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class DeleteProductView(APIView):
  permission_classes = [IsAuthenticated]

  def delete(self, request, pk, format=None):
    product = get_object_or_404(Product, pk=pk, deleted_at__isnull=True)
    product.deleted_at = timezone.now()
    product.updated_by = request.user
    product.save()

    return Response({'message': 'Product deleted successfully'}, status=status.HTTP_204_NO_CONTENT)

class DeleteProductsByIdsView(APIView):
  permission_classes = [IsAuthenticated]

  def post(self, request):
    try:
      product_ids = request.data if isinstance(request.data, list) else request.data.get('ids', [])

      if not product_ids:
        return Response(
          {'error': 'No se proporcionaron IDs de productos'},
          status=status.HTTP_400_BAD_REQUEST
        )

      # Convertir a enteros
      try:
        product_ids = [int(id) for id in product_ids]
      except (ValueError, TypeError):
        return Response(
          {'error': 'IDs de productos no válidos'},
          status=status.HTTP_400_BAD_REQUEST
        )

      # Productos existentes (no eliminados lógicamente)
      existing_products = Product.objects.filter(
        id__in=product_ids,
        deleted_at__isnull=True
      )

      # Validación: si TODOS los IDs existen
      if existing_products.count() == len(product_ids):
        # Eliminar físicamente
        deleted_count, _ = existing_products.delete()
        return Response({
          'message': f'{deleted_count} productos eliminados físicamente',
          'deleted_count': deleted_count,
          'mode': 'hard_delete'
        }, status=status.HTTP_200_OK)
      else:
        # Eliminar lógicamente
        updated = existing_products.update(
          deleted_at=timezone.now(),
          updated_by=request.user
        )
        return Response({
          'message': f'{updated} productos marcados como eliminados',
          'deleted_count': updated,
          'mode': 'soft_delete'
        }, status=status.HTTP_200_OK)

    except Exception as e:
      traceback.print_exc()
      return Response(
        {'error': f'Error al eliminar productos: {str(e)}'},
        status=status.HTTP_500_INTERNAL_SERVER_ERROR
      )

class ProductCategoriesView(APIView):
  permission_classes = [IsAuthenticated]
  parser_classes = [JSONParser, FormParser, MultiPartParser]

  def get(self, request, product_id, format=None):
    product = get_object_or_404(Product, pk=product_id, deleted_at__isnull=True)
    categories = product.product_categories.all()
    serializer = ProductCategorySerializer(categories, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

  def post(self, request, product_id, format=None):
    product = get_object_or_404(Product, pk=product_id, deleted_at__isnull=True)
    category_id = request.data.get('category_id')
    
    if not category_id:
      return Response(
        {'error': 'category_id is required'},
        status=status.HTTP_400_BAD_REQUEST
      )

    # Check if relationship already exists
    if product.categories.filter(id=category_id).exists():
      return Response(
        {'error': 'This category is already assigned to the product'},
        status=status.HTTP_400_BAD_REQUEST
      )

    product_category = ProductCategory.objects.create(
      product=product,
      category_id=category_id
    )
    
    serializer = ProductCategorySerializer(product_category)
    return Response(serializer.data, status=status.HTTP_201_CREATED)

  def delete(self, request, product_id, format=None):
    product = get_object_or_404(Product, pk=product_id, deleted_at__isnull=True)
    category_id = request.data.get('category_id')
    
    if not category_id:
      return Response(
        {'error': 'category_id is required'},
        status=status.HTTP_400_BAD_REQUEST
      )

    product_category = get_object_or_404(
      ProductCategory,
      product=product,
      category_id=category_id
    )
    
    product_category.delete()
    return Response(
      {'message': 'Category removed from product successfully'},
      status=status.HTTP_204_NO_CONTENT
    )
    
class ExportProductsView(APIView):
  permission_classes = [IsAuthenticated]

  def post(self, request):
    # Decodificar el cuerpo de la solicitud para obtener los IDs
    try:
      body = json.loads(request.body)
      product_ids = body.get('ids', [])
    except json.JSONDecodeError:
      return JsonResponse({'error': 'Invalid input'}, status=400)

    # Obtener los productos según los IDs proporcionados
    try:
      if product_ids:
        products = Product.objects.filter(id__in=product_ids)
      else:
        products = Product.objects.filter(deleted_at__isnull=True)
    except Exception as e:
      return JsonResponse(
        {'error': 'Failed to fetch products'}, 
        status=500
      )

    # Crear un archivo Excel en memoria
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"

    # Agregar encabezados
    headers = ["Codigo", "Nombre", "Precio(CF)", "Precio(SF)", "Precio(Caja)", "Ctd. en caja", "Unidad de medida"]
    sheet.append(headers)

    # Estilo para los encabezados
    for cell in sheet[1]:
      cell.font = Font(bold=True)

    # Variables para ajustar el ancho de las columnas
    column_widths = [len(header) for header in headers]

    # Agregar datos de los productos
    for product in products:
      # Manejar campos posibles None
      sku = product.sku if product.sku else ""
      
      pcf = product.featured_pcf if product.featured_pcf is not None else 0.0
      psf = product.featured_psf if product.featured_psf is not None else 0.0
      pbox = product.featured_pbox if product.featured_pbox is not None else 0.0
      qtBox = product.quantity_in_box if product.quantity_in_box is not None else 0
      unit = product.unit_of_measurement.name if product.unit_of_measurement else ""

      row = [
        sku,
        product.name,
        f"{pcf:.2f}",
        f"{psf:.2f}",
        f"{pbox:.2f}",
        f"{qtBox:.2f}",
        # f"{product.cost:.2f}",
        unit
      ]

      sheet.append(row)

      # Actualizar anchos máximos de columnas
      for i, value in enumerate(row):
        column_widths[i] = max(column_widths[i], len(str(value)))

    # Ajustar el ancho de las columnas
    for i, width in enumerate(column_widths):
      # Asegurar un ancho mínimo razonable
      adjusted_width = max(width, 10) + 2  # Agregar un pequeño margen
      sheet.column_dimensions[get_column_letter(i+1)].width = adjusted_width

    # Guardar el libro de trabajo en el buffer
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(
      output.getvalue(),
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products-export.xlsx'
    return response

class ImportProductsView(APIView):
    
  permission_classes = [IsAuthenticated]
  parser_classes = [MultiPartParser, FormParser]
  
  def post(self, request):
    if 'file' not in request.FILES:
      return JsonResponse({'error': 'No file provided'}, status=400)
        
    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name
    user_id = request.user.id
    
    try:
      if file_name.endswith('.xlsx'):
        products_data = parse_excel(uploaded_file, user_id)
      elif file_name.endswith('.csv'):
        products_data = parse_csv(uploaded_file)
      else:
        return JsonResponse(
          {'error': 'Unsupported file format. Only .xlsx and .csv are supported'}, 
          status=400
        )
      
      # Cargar unidades de medida
      units = UnitOfMeasurement.objects.all()
      shortcut_map = {u.shortcut.strip().upper(): u.id for u in units if u.shortcut}
      name_map = {u.name.strip().upper(): u.id for u in units if u.name}
        
      # Obtener todos los SKUs y nombres ya registrados en BD
      existing_products = Product.objects.values_list("sku", "name")
      existing_skus = {sku for sku, _ in existing_products if sku}
      existing_names = {name.lower() for _, name in existing_products if name}
      
      # Filtrar productos duplicados
      new_products = []
      skipped_products = []
      restored_products = []

      for product in products_data:
        sku = product.get("sku")
        name = product.get("name")
        unit_value = product.get("unit")

        # Buscar producto existente por sku o name
        existing = Product.objects.filter(
          models.Q(sku=sku) | models.Q(name__iexact=name)
        ).first()

        # Si ya existe y está eliminado → restaurarlo
        if existing and existing.deleted_at is not None:
          unit_id = None
          if unit_value:
            unit_value = str(unit_value).strip().upper()
            unit_id = shortcut_map.get(unit_value)

          existing.deleted_at = None
          existing.updated_at = timezone.now()
          existing.updated_by_id = user_id
          existing.cost = product['cost']
          existing.unit_of_measurement_id = unit_id
          existing.save()
          restored_products.append(name)
          continue
        
        # Si ya existe y no está eliminado → saltarlo
        if existing and existing.deleted_at is None:
          skipped_products.append(name)
          continue

        # Detectar unidad de medida
        unit_id = None
        if unit_value:
          unit_value = str(unit_value).strip().upper()
          unit_id = shortcut_map.get(unit_value)
        
        new_products.append(
          Product(
            sku=sku,
            name=name,
            unit_of_measurement_id=unit_id,
            prices_cf=product.get('prices_cf', {}),
            prices_sf=product.get('prices_sf', {}),
            prices_box=product.get('prices_box', {}),
            featured_pcf=product.get('featured_pcf'),
            featured_psf=product.get('featured_psf'),
            featured_pbox=product.get('featured_pbox'),
            cost=product['cost'],
            created_by_id=user_id
          )
        )

      # Crear solo los que son nuevos
      if new_products:
        Product.objects.bulk_create(new_products)
      
      return JsonResponse({
        'message': f'Importación finalizada. {len(new_products)} nuevo(s), {len(restored_products)} restaurado(s), {len(skipped_products)} omitido(s).',
        'restored': restored_products,
        'skipped': skipped_products  # Para que sepas cuáles se omitieron
      })

    except Exception as e:
      return JsonResponse({'error': str(e)}, status=400)